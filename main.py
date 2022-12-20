import openpyxl
import time

start_time = time.time()

wb = openpyxl.load_workbook(filename='Dataset.xlsx')
sheet = wb['Sheet1']

social_networks = ['facebook.com', 'instagram.com', 'twiter.com', 'vk.com',
                   'viber.com', 'telegram.org', 'wechat.com', 'ok.ru',
                   'my.mail.ru', 'tambrl.com', 'vkrugudruzei.ru', 'nsportal.ru']

video_hostings = ['youtube.com', 'tiktok.com', 'twitch.tv', 'likee.video',
                  'snapchat.com', 'subscribe.ru', 'dailymotion.com', 'vimeo.com',
                  'rutube.ru', 'sproutvideo.com', 'vidyard.com', 'videos.kaltura.com', 'blogger.com']

dating_websites = ['badoo.com', 'tinder.com', 'pinterest.ru', 'linkedin.com',
                   'clubhouse.com', 'toxicbun.com', 'vimple.ru', 'dacast.com',
                   'wistia.com', 'twentythree.com', 'jwplayer.ru', 'teamo.ru']

informational_resources = ['livejournal.com', 'cincopa.com', 'brightcove.com', 'panopto.com',
                           'liveinternet.ru', 'habr.com', 'currents.google.com', 'last.fm',
                           'whotrades.com', 'povarenok.ru', 'scipeople.ru', 'agrobook.ru', 'fotokto.ru']

print("Какую функцию сделать (Рассчитать K-анонимити/Обезличить данные)?")
a = input()
if a == "Рассчитать K-анонимити":
    print("Введите количество квази-идентификаторов: ")
    n = int(input())

    ind = []
    for i in range(n):
        print("Введите номер квази-идентификатора: ")
        ind.append(int(input()))
    rows = {}
    st = []
    rows.clear()
    for i in range(1, 260001):
        st.clear()
        s = ""
        for j in range(n):
            st.append(sheet.cell(row=i, column=ind[j]).value)
        s = str(st)
        if s in rows:
            rows[s] += 1
        else:
            rows[s] = 1
    K = 260001

    for i in rows:
        if rows[i] < K:
            K = rows[i]
    print("К-анонимити ", K)
    wb.close()

elif a == "Обезличить данные":
    for i in range(2, 260002):
        sheet['A' + str(i - 1)] = "XXX.XXX.XXX.XXX"

        platform = sheet['C' + str(i)].value
        if platform in social_networks:
            sheet['B' + str(i - 1)] = "Соц.сеть"
        elif platform in video_hostings:
            sheet['B' + str(i - 1)] = "Видеохостинг"
        elif platform in dating_websites:
            sheet['B' + str(i - 1)] = "Сайт знакомств"
        else:
            sheet['B' + str(i - 1)] = "Инфоресурс"

        date = sheet['D' + str(i)].value
        date = date.split("/")
        if int(date[1]) == 12 or int(date[1]) == 1 or int(date[1]) == 2:
            date = "Зима"
            sheet['C' + str(i - 1)] = date
        elif int(date[1]) == 3 or int(date[1]) == 4 or int(date[1]) == 5:
            date = "Весна"
            sheet['C' + str(i - 1)] = date
        elif int(date[1]) == 6 or int(date[1]) == 7 or int(date[1]) == 8:
            date = "Лето"
            sheet['C' + str(i - 1)] = date
        else:
            date = "Осень"
            sheet['C' + str(i - 1)] = date

        ad_time = sheet['F' + str(i)].value
        ad_time = ad_time.split()
        ad_time = ad_time[0].split(":")
        sheet['D' + str(i - 1)] = ad_time[0]

        product = sheet['G' + str(i)].value
        product = product.split()
        sheet['E' + str(i - 1)] = product[0]

        sheet['F' + str(i - 1)] = ""
        sheet['G' + str(i - 1)] = ""

    for i in range(1, 8):
        sheet.cell(row=260001, column=i).value = ""

    wb.save('Dataset1.xlsx')
    wb.close()

    wb = openpyxl.load_workbook(filename='Dataset1.xlsx')
    sheet = wb['Sheet1']

    print("Введите количество квази-идентификаторов: ")
    n = int(input())

    ind = []
    for i in range(n):
        print("Введите номер квази-идентификатора: ")
        ind.append(int(input()))

    st = []
    rows = {}
    for i in range(1, 260001):
        st.clear()
        s = ""
        for j in range(n):
            st.append(sheet.cell(row=i, column=ind[j]).value)
        s = str(st)
        if s in rows:
            rows[s] += 1
        else:
            rows[s] = 1

    unique = []
    bad = []
    bad_excep = []
    for i in rows:
        if rows[i] == 1:
            unique.append(str(i))
        if rows[i] < 5:
            bad.append(str(rows[i]) + " " + str(i))
            bad_excep.append(str(i))
    if len(bad) == 0:
        print("Плохих значений нет")
    elif len(bad) < 5:
        print("Все плохие значения:")
        for i in range(len(bad)):
            print(bad[i])
    else:
        print("5 плохих значений:")
        for i in range(5):
            print(bad[i])
    print(f"Процент плохих строк в наборе: {(len(bad) / 260000 * 100):.2f}", "%")
    print("Количество уникальных строк: ", len(unique))
    if len(unique) != 0:
        print("Уникальные строки: ")
        for i in unique:
            print(i)

    w = 1
    for i in range(1, 260001):
        s = ""
        st.clear()
        for j in range(n):
            st.append(sheet.cell(row=i, column=ind[j]).value)
        s = str(st)
        if s not in bad_excep:
            for p in range(1, 6):
                sheet.cell(row=w, column=p).value = sheet.cell(row=i, column=p).value
            w += 1

    for i in range(w, 260001):
        for p in range(1, 6):
            sheet.cell(row=i, column=p).value = ""
    wb.save('Dataset2.xlsx')
    wb.close()

    wb = openpyxl.load_workbook(filename='Dataset2.xlsx')
    sheet = wb['Sheet1']

    rows.clear()
    for i in range(1, w):
        st.clear()
        s = ""
        for j in range(n):
            st.append(sheet.cell(row=i, column=ind[j]).value)
        s = str(st)
        if s in rows:
            rows[s] += 1
        else:
            rows[s] = 1

    minimum = w

    for i in rows:
        if rows[i] < minimum:
            minimum = rows[i]
    wb.close()
    print("К-анонимити ", minimum)
    print("--- %s seconds ---" % (time.time() - start_time))
else:
    print("Ошибка ввода")
